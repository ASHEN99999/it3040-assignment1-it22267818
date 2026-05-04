import openpyxl

wb = openpyxl.load_workbook('it22267818.xlsx')
ws = wb.active
print('Sheet name:', ws.title)
print('Max row:', ws.max_row)
print('Max col:', ws.max_column)
for row in ws.iter_rows(values_only=True):
    print(row)